"""
Export router: downloadable summary reports of sources.

Produces an Excel workbook with one row per source (title, type, file/URL,
topics, AI-generated summary from insights, notebooks, dates) so a folder of
ingested documents can be reviewed at a glance and revisited later.
"""

import io
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from open_notebook.database.repository import ensure_record_id, repo_query

router = APIRouter(prefix="/export", tags=["export"])

# Characters not allowed in xlsx cell values (control chars except \t \n \r)
_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

SUMMARY_EXCERPT_CHARS = 500


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return _ILLEGAL_XLSX_CHARS.sub("", str(value))


def _source_type(asset: Optional[Dict[str, Any]]) -> str:
    if asset and asset.get("url"):
        return "Link"
    if asset and asset.get("file_path"):
        return "File"
    return "Text"


def _source_location(asset: Optional[Dict[str, Any]]) -> str:
    if asset and asset.get("url"):
        return str(asset["url"])
    if asset and asset.get("file_path"):
        return Path(str(asset["file_path"])).name
    return ""


def _summarize(source: Dict[str, Any], insights: List[Dict[str, Any]]) -> str:
    """Best available summary: summary-type insights, then any insights,
    then an excerpt of the extracted text."""
    summaries = [
        i["content"]
        for i in insights
        if "summary" in str(i.get("insight_type", "")).lower() and i.get("content")
    ]
    if summaries:
        return "\n\n".join(summaries)

    other = [
        f"{i.get('insight_type', 'Insight')}: {i['content']}"
        for i in insights
        if i.get("content")
    ]
    if other:
        return "\n\n".join(other)

    full_text = source.get("full_text") or ""
    if full_text.strip():
        excerpt = full_text.strip()[:SUMMARY_EXCERPT_CHARS]
        suffix = "…" if len(full_text.strip()) > SUMMARY_EXCERPT_CHARS else ""
        return f"(excerpt) {excerpt}{suffix}"

    return ""


@router.get("/summary-report")
async def export_summary_report(
    notebook_id: Optional[str] = Query(
        default=None, description="Limit the report to one notebook"
    ),
):
    """Build and stream an .xlsx summarizing every source."""
    try:
        report_scope = "all-sources"
        if notebook_id:
            nb_rows = await repo_query(
                "SELECT name FROM ONLY $nb",
                {"nb": ensure_record_id(notebook_id)},
            )
            if not nb_rows:
                raise HTTPException(status_code=404, detail="Notebook not found")
            nb_name = (
                nb_rows.get("name")
                if isinstance(nb_rows, dict)
                else nb_rows[0].get("name")
            ) or "notebook"
            report_scope = re.sub(r"[^A-Za-z0-9._-]+", "-", nb_name).strip("-") or "notebook"

            rows = await repo_query(
                """
                SELECT * FROM (
                    SELECT in AS source FROM reference WHERE out=$nb FETCH source
                ) ORDER BY source.updated DESC
                """,
                {"nb": ensure_record_id(notebook_id)},
            )
            sources = [r["source"] for r in rows if r.get("source")]
        else:
            sources = await repo_query("SELECT * FROM source ORDER BY updated DESC")

        # One query each for insights and notebook links, grouped in Python,
        # to avoid per-source round trips on bulky folders.
        insight_rows = await repo_query(
            "SELECT source, insight_type, content FROM source_insight"
        )
        insights_by_source: Dict[str, List[Dict[str, Any]]] = {}
        for row in insight_rows or []:
            insights_by_source.setdefault(str(row.get("source")), []).append(row)

        ref_rows = await repo_query(
            "SELECT in AS source, out.name AS notebook FROM reference"
        )
        notebooks_by_source: Dict[str, List[str]] = {}
        for row in ref_rows or []:
            if row.get("notebook"):
                notebooks_by_source.setdefault(str(row.get("source")), []).append(
                    str(row["notebook"])
                )

        wb = Workbook()
        ws = wb.active
        ws.title = "Sources"

        headers = [
            "Title",
            "Type",
            "File / URL",
            "Topics",
            "Summary",
            "Notebooks",
            "Added",
            "Updated",
        ]
        widths = [40, 8, 45, 30, 90, 25, 20, 20]
        for col, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A2"

        wrap = Alignment(wrap_text=True, vertical="top")
        for row_idx, source in enumerate(sources or [], start=2):
            source_id = str(source.get("id"))
            insights = insights_by_source.get(source_id, [])
            values = [
                _clean(source.get("title") or "Untitled source"),
                _source_type(source.get("asset")),
                _clean(_source_location(source.get("asset"))),
                _clean(", ".join(source.get("topics") or [])),
                _clean(_summarize(source, insights)),
                _clean(", ".join(notebooks_by_source.get(source_id, []))),
                _clean(str(source.get("created") or "")[:19]),
                _clean(str(source.get("updated") or "")[:19]),
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = wrap

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        filename = f"summary-report-{report_scope}-{timestamp}.xlsx"
        logger.info(
            f"Summary report generated: {len(sources or [])} sources ({filename})"
        )
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating summary report: {str(e)}")
        logger.exception(e)
        raise HTTPException(
            status_code=500, detail=f"Error generating summary report: {str(e)}"
        )
